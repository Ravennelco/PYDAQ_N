import nidaqmx
import time 
from numpy import arange
from openpyxl import Workbook
from tkinter import *
from threading import Thread


class Testing():
    def __init__(self,analog=None):
        self.task = nidaqmx.Task()
        self.task2 = nidaqmx.Task()
        self.task3 = nidaqmx.Task()
        self.value = [0,0,0,0]
        self.hilo = None
        self.runS = False
        self.wb = Workbook()
        self.ws = self.wb.active
        self.r = 0
        #self.DAQ()

    def HiloSerial(self):
        if self.hilo == None:
            self.hilo=Thread(target=self.Excel)
            self.task.ai_channels.add_ai_voltage_chan('Dev1/ai0',min_val= 0, max_val = 5)
            self.task.ai_channels.add_ai_voltage_chan('Dev1/ai1',min_val= 0, max_val = 5)
            self.task.ai_channels.add_ai_voltage_chan('Dev1/ai2',min_val= 0, max_val = 5)
            self.task.ai_channels.add_ai_voltage_chan('Dev1/ai3',min_val= 0, max_val = 5)
            self.task3.ao_channels.add_ao_voltage_chan('Dev1/ao0','mychanel',0,3.3)
            self.task2.ao_channels.add_ao_voltage_chan('Dev1/ao1','mychanel',0,3.3)
            self.hilo.start()


    def Excel(self,a=False,x=None,y=None):
        self.runS = a
        count = 1
        self.value_a = [0,0,0,0]
        arr = [0,0,0,0]
        c=0
        self.value = self.task.read()
        self.r += 1 
        print(x)
        while (self.runS) and (count < 100):
            #time.sleep(0.1)
            self.value = self.task.read()
            #self.task2.write(2.5)
            self.task2.write((2.49E-16 + 2.8*int(x) + 9.79E-14*int(x)**2 + -3.77E-13*int(x)**3 + 6.86E-13*int(x)**4 + -5.88E-13*int(x)**5 + 1.91E-13*int(x)**6)/100)
            self.task3.write((2.49E-16 + 2.8*int(y) + 9.79E-14*int(y)**2 + -3.77E-13*int(y)**3 + 6.86E-13*int(y)**4 + -5.88E-13*int(y)**5 + 1.91E-13*int(y)**6)/100)
            #self.task2.write((2.49E-16 + 2.8*int(c) + 9.79E-14*int(c)**2 + -3.77E-13*int(c)**3 + 6.86E-13*int(c)**4 + -5.88E-13*int(c)**5 + 1.91E-13*int(c)**6)/100)
            #print(self.value[2]+0.07)    
            count +=1
            c +=1
            self.value_a[0] = self.value[0] + self.value_a[0]
            self.value_a[1] = self.value[1] + self.value_a[1]
            self.value_a[2] = self.value[2] + self.value_a[2]
            self.value_a[3] = self.value[3] + self.value_a[3]
            #print(self.value[3])
            #self.celdaI = f'A{self.count}'
            #self.celdaD = f'B{self.count}'
            #self.ws[self.celdaI] = self.value[0]
            #self.ws[self.celdaD] = self.value[1] 
        arr[0] = (self.value_a[0])/count #Galga 1
        arr[1] = (self.value_a[1])/count #Galga 2 
        arr[2] = (self.value_a[2])/count #Corriente 1
        arr[3] = (self.value_a[3])/count  #Corriente 2
        self.i1 = f'A{self.r}'
        self.i2 = f'B{self.r}'
        self.F1 = f'C{self.r}'
        self.F2 = f'D{self.r}'
        F1 = 25000*arr[0] - 0.025
        F2 = 25000*arr[1] - 0.025
        I1 = -9668.5*(arr[2])**6 + 11067*(arr[2])**5 -4938.8*(arr[2])**4 + 1078.1*(arr[2])**3 -117.83*(arr[2])**2 + 8.7677*(arr[2]) - 0.0427#porfavor pruebe la función
        I2 = -9668.5*(arr[3])**6 + 11067*(arr[3])**5 -4938.8*(arr[3])**4 + 1078.1*(arr[3])**3 -117.83*(arr[3])**2 + 8.7677*(arr[3]) - 0.0427#porfavor pruebe la función
        self.ws[self.i1] = "{:.6f}".format(I1)
        self.ws[self.i2] = "{:.6f}".format(I2)
        self.ws[self.F1] = "{:.6f}".format(F1)
        self.ws[self.F2] = "{:.6f}".format(F2)      
        print(f"\t CELDA 1   CELDA 2   i1   i2 \n \
        {arr[0]:.6f}   {arr[1]:.6f}  {arr[2]:.6f} {arr[3]:.6f}")
        self.wb.save('calibra.xlsx')
    
    def Salir(self):
        self.runS = False
        self.hilo.join()
        self.task.stop() 
        self.task.close()
        print("Terminoooo!!!") 
    
 


class Ventana(Frame,Testing):
    def __init__(self,master=None,analog=None):
        Frame.__init__(self,master=None)
        Testing.__init__(self,analog=None)
        self.master = master
        self.analog = analog
        self.widged()


    def crear_marco_sup(self):
        mac_sup = Frame(self.master)
        mac_sup.config(height = 30)
        mac_sup.grid(row=0, columnspan=12, rowspan=10, padx= 5, pady = 5)
        Label(mac_sup, text = "ciclo de trabajo").grid(row=0, column=0)
        self.ciclo = StringVar()
        self.c_val = StringVar(value=1)
        self.c2_val = StringVar(value=1)
        #print(self.ciclo.get())
        Entry(mac_sup, textvariable=self.ciclo).grid(row=0, column=1, padx=7, pady=2)
        Label(mac_sup, text='Bobina 1:').grid(row=0, column=2)
        self.baud = Spinbox(mac_sup, from_=0, to=100, width=7, values=arange(0,101),textvariable=self.c_val)
        self.baud.grid(row=0, column=3)
        Label(mac_sup, text='Bobina 2:').grid(row=1, column=2)
        self.baud2 = Spinbox(mac_sup, from_=0, to=100, width=7, values=arange(0,101),textvariable=self.c2_val)
        self.baud2.grid(row=1, column=3)
        



    def registrar(self):
        self.analog.Excel(True,self.c_val.get(),self.c2_val.get())
        
        #self.analog.Excel(True,self.ciclo.get())

    

    def Matar(self):
        self.analog.Salir()
        self.master.destroy()


    
    def menu(self):
        self.crear_marco_sup()
        B1 = Button(self.master,text = "Registrar",command=self.registrar)
        B1.place(x = 30,y = 30)
        B2 = Button(self.master,text = "Salir",command=self.Matar)
        B2.place(x = 120,y = 30)


    def widged(self):
        self.menu()
        self.master.mainloop()



if __name__ == '__main__':
    root = Tk() # instancia del canvas de tkinter
    ex = Testing()
    ex.HiloSerial()
    root.geometry("800x600")
    root.title("Calibracion")
    app = Ventana(root,ex)

   #testing()
