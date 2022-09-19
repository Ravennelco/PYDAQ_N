import nidaqmx
import time 
from openpyxl import Workbook
'''
task = nidaqmx.Task()
task.ao_channels.add_ao_voltage_chan('Dev2/ao0','mychanel',0,5)
task.start()

value = 2
task.write(value)
task.stop()
task.close()
'''
wb = Workbook()
ws = wb.active
    

task = nidaqmx.Task()
task2 = nidaqmx.Task()
task3 = nidaqmx.Task()
task.ai_channels.add_ai_voltage_chan('Dev1/ai0',min_val= 0, max_val = 6)
task.ai_channels.add_ai_voltage_chan('Dev1/ai1',min_val= 0, max_val = 6)
task.ai_channels.add_ai_voltage_chan('Dev1/ai2',min_val= 0, max_val = 6)
task.ai_channels.add_ai_voltage_chan('Dev1/ai3',min_val= 0, max_val = 6)
task3.ao_channels.add_ao_voltage_chan('Dev1/ao0','mychanel',0,3.3)
task2.ao_channels.add_ao_voltage_chan('Dev1/ao1','mychanel',0,3.3)
#task.write(value)
#time.sleep(2)
value2 = 0
value = [0,0,0]
for count in range(1,10001,1):
    #value = value + 0.03
    time.sleep(0.1)
    #print(count)
    value = task.read()
    #task2.write(0.105)
    #task3.write(0.2)
    print(f'{value[2]} {value[3]}')
    #celdaI = f'A{count}'
    #celdaD = f'B{count}'
    #ws[celdaI] = value[0]
    #ws[celdaD] = value[1]
#wb.save('Datos.xlsx')

#value = 0.5
#task.write(value)
task.stop()
task.close()

